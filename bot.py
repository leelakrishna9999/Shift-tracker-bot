# Shift-tracker-bot
import logging
import re
import math
import requests
from datetime import datetime, timedelta
from telegram import Update
from telegram.ext import Application, MessageHandler, filters, ContextTypes
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

BOT_TOKEN           = “8792621305:AAFu0aKuak02DgPOoEwp3sLgpaRrmx4CMVU”
YOUR_USER_ID        = 8293531859
YOUR_NAME           = “Leelakrishna”
EXCEL_FILE          = “WorkHoursTracker.xlsx”
HOME_ADDRESS        = “16 Oranmore Street, Belfast, BT13 2RU”
HOME_LAT            = 54.5973
HOME_LNG            = -5.9301
BUS_SPEED_KMH       = 20
BUS_WAIT_MINS       = 10
WALK_TO_STOP_MINS   = 5
LEAVE_BUFFER_MINS   = 15
WAKE_UP_BUFFER_MINS = 90

PAY_RATES_PRE_APRIL = {
“weekday_day”:   12.50,
“weekday_night”: 13.00,
“saturday”:      13.00,
“sunday”:        15.00,
“bank_holiday”:  15.00,
“overtime”:      13.00,
}
PAY_RATES_APRIL_ONWARDS = {
“weekday_day”:   13.00,
“weekday_night”: 14.00,
“saturday”:      14.00,
“sunday”:        16.00,
“bank_holiday”:  18.00,
“overtime”:      15.00,
}
APRIL_2026 = datetime(2026, 4, 1)

def get_active_rates(shift_date):
if shift_date < APRIL_2026:
return PAY_RATES_PRE_APRIL, “Feb/Mar 2026”
return PAY_RATES_APRIL_ONWARDS, “April 2026+”

BANK_HOLIDAYS = {
“2026-01-01”,“2026-03-17”,“2026-04-10”,“2026-04-13”,
“2026-05-04”,“2026-05-25”,“2026-07-12”,“2026-08-31”,
“2026-12-25”,“2026-12-28”
}

SHIFT_TYPES = {
“long day”:   {“start”: 8,  “end”: 20, “break”: 1.0},
“night duty”: {“start”: 20, “end”: 8,  “break”: 1.0},
“night”:      {“start”: 20, “end”: 8,  “break”: 1.0},
“early”:      {“start”: 6,  “end”: 14, “break”: 1.0},
“late”:       {“start”: 14, “end”: 22, “break”: 1.0},
“short day”:  {“start”: 8,  “end”: 16, “break”: 0.5},
}

logging.basicConfig(format=”%(asctime)s - %(levelname)s - %(message)s”, level=logging.INFO)

def geocode(address):
try:
url = “https://nominatim.openstreetmap.org/search”
params = {“q”: address, “format”: “json”, “limit”: 1, “countrycodes”: “gb”}
headers = {“User-Agent”: “LeelakrishnaShiftBot/1.0”}
r = requests.get(url, params=params, headers=headers, timeout=10)
data = r.json()
if data:
return float(data[0][“lat”]), float(data[0][“lon”])
except:
pass
return None, None

def distance_km(lat1, lng1, lat2, lng2):
R = 6371
dlat = math.radians(lat2 - lat1)
dlng = math.radians(lng2 - lng1)
a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlng/2)**2
return round(R * 2 * math.asin(math.sqrt(a)) * 1.3, 1)

def travel_time(dist_km):
return WALK_TO_STOP_MINS + BUS_WAIT_MINS + int((dist_km / BUS_SPEED_KMH) * 60)

def get_schedule(shift_date, start_hr, travel_mins):
shift_start = shift_date.replace(hour=start_hr, minute=0, second=0, microsecond=0)
leave_home  = shift_start - timedelta(minutes=travel_mins + LEAVE_BUFFER_MINS)
wake_up     = leave_home  - timedelta(minutes=WAKE_UP_BUFFER_MINS)
return wake_up.strftime(”%H:%M”), leave_home.strftime(”%H:%M”), shift_start.strftime(”%H:%M”)

def get_pay_rate(date_obj, shift_type):
rates, period = get_active_rates(date_obj)
date_str = date_obj.strftime(”%Y-%m-%d”)
if date_str in BANK_HOLIDAYS:
return rates[“bank_holiday”], “Bank Holiday”, period
wd = date_obj.weekday()
if wd == 6: return rates[“sunday”],        “Sunday”,        period
if wd == 5: return rates[“saturday”],      “Saturday”,      period
if “night” in shift_type.lower():
return rates[“weekday_night”], “Weekday Night”, period
return rates[“weekday_day”], “Weekday Day”, period

def shift_hours(shift_type):
for key, val in SHIFT_TYPES.items():
if key in shift_type.lower():
raw = val[“end”] - val[“start”]
if raw <= 0: raw += 24
return raw - val[“break”], val[“start”], val[“end”], val[“break”]
return 11, 8, 20, 1.0

def parse_shift(text):
months = {“january”:1,“february”:2,“march”:3,“april”:4,“may”:5,“june”:6,
“july”:7,“august”:8,“september”:9,“october”:10,“november”:11,“december”:12}
result = {}
dm = re.search(r’(\d{1,2})(?:st|nd|rd|th)?\s+of\s+(\w+)’, text, re.I) or   
re.search(r’(\d{1,2})(?:st|nd|rd|th)?\s+(\w+)’, text, re.I)
if dm:
day = int(dm.group(1))
month = months.get(dm.group(2).lower(), datetime.now().month)
try:    result[“date”] = datetime(2026, month, day)
except: result[“date”] = datetime.now()
else:
result[“date”] = datetime.now()

```
lines = [l.strip() for l in text.split('\n') if l.strip()]
result["location"] = "Unknown Location"
for i, line in enumerate(lines):
    if re.search(r'shift', line, re.I) and ':' in line and i+1 < len(lines):
        result["location"] = lines[i+1]
        break

pc = re.search(r'\b[A-Z]{1,2}\d{1,2}\s?\d[A-Z]{2}\b', text, re.I)
result["postcode"] = pc.group(0).upper() if pc else ""
result["geo_query"] = (
    f"{result['location']}, {result['postcode']}, Northern Ireland"
    if result["postcode"] else
    f"{result['location']}, Belfast, Northern Ireland"
)

um = re.search(r'(\w+)\s+[Uu]nit', text)
result["unit"] = um.group(0).title() if um else "Not specified"

result["shift_type"] = "Long Day"
if "long day" in text.lower():
    result["shift_type"] = "Long Day"
else:
    for key in SHIFT_TYPES:
        if key in text.lower():
            result["shift_type"] = key.title()
            break
return result
```

DARK=“0D2137”; TEAL=“0E7C86”; GREEN=“1A6B3C”; WHITE=“FFFFFF”; LGREY=“F5F6F8”; LGRN=“F0FAF3”
HEADER_COLS=[”#”,“Date”,“Day”,“Location”,“Unit”,“Shift Type”,“Day Type”,“Start”,“End”,“Break(hrs)”,“Hours”,“Rate(£)”,“Earnings(£)”,“Distance(km)”,“Bus(mins)”,“Wake Up”,“Leave Home”]
COL_WIDTHS=[5,13,12,32,14,12,18,8,8,10,8,9,12,13,11,10,12]

def thin_border(color=“CCCCCC”):
s=Side(style=“thin”,color=color)
return Border(left=s,right=s,top=s,bottom=s)

def get_or_create_month_sheet(wb, date_obj):
sheet_name = date_obj.strftime(”%B %Y”)
if sheet_name in wb.sheetnames:
return wb[sheet_name]
ws = wb.create_sheet(title=sheet_name)
ws.sheet_view.showGridLines = False
ws.merge_cells(f”A1:{get_column_letter(len(HEADER_COLS))}1”)
t=ws[“A1”]; t.value=f”LEELAKRISHNA VALLURI — SHIFTS — {sheet_name.upper()}”
t.font=Font(name=“Arial”,bold=True,size=12,color=WHITE)
t.fill=PatternFill(“solid”,fgColor=DARK)
t.alignment=Alignment(horizontal=“center”,vertical=“center”)
ws.row_dimensions[1].height=28
ws.merge_cells(f”A2:{get_column_letter(len(HEADER_COLS))}2”)
s=ws[“A2”]; s.value=“CarrikCare Group | Band 3 Senior Care Assistant | SE-028”
s.font=Font(name=“Arial”,size=10,color=“AAAAAA”,italic=True)
s.fill=PatternFill(“solid”,fgColor=DARK)
s.alignment=Alignment(horizontal=“center”,vertical=“center”)
ws.row_dimensions[2].height=18; ws.row_dimensions[3].height=6
for col,(h,w) in enumerate(zip(HEADER_COLS,COL_WIDTHS),1):
c=ws.cell(row=4,column=col,value=h)
c.font=Font(name=“Arial”,bold=True,size=10,color=WHITE)
c.fill=PatternFill(“solid”,fgColor=TEAL)
c.alignment=Alignment(horizontal=“center”,vertical=“center”,wrap_text=True)
c.border=thin_border(WHITE)
ws.column_dimensions[get_column_letter(col)].width=w
ws.row_dimensions[4].height=28
ws.freeze_panes=“A5”
return ws

def sort_sheet_by_date(ws):
rows_data=[]
for row in ws.iter_rows(min_row=5,values_only=True):
if row[0] is not None:
rows_data.append(list(row))
if not rows_data: return
def parse_date(row):
try: return datetime.strptime(str(row[1]),”%d/%m/%Y”)
except: return datetime.min
rows_data.sort(key=parse_date)
for row in ws.iter_rows(min_row=5,max_row=ws.max_row):
for cell in row:
cell.value=None
cell.fill=PatternFill(“solid”,fgColor=WHITE)
for i,row_data in enumerate(rows_data):
actual_row=5+i; bg=LGRN if i%2==0 else WHITE; row_data[0]=i+1
for col,val in enumerate(row_data,1):
c=ws.cell(row=actual_row,column=col,value=val)
c.fill=PatternFill(“solid”,fgColor=bg)
c.font=Font(name=“Arial”,size=10)
c.alignment=Alignment(horizontal=“center”,vertical=“center”)
c.border=thin_border()
ws.row_dimensions[actual_row].height=18

def get_or_create_summary(wb):
if “Summary” in wb.sheetnames: return wb[“Summary”]
ws=wb.create_sheet(title=“Summary”,index=0)
ws.sheet_view.showGridLines=False
ws.merge_cells(“A1:G1”); t=ws[“A1”]
t.value=“MONTHLY SUMMARY — LEELAKRISHNA VALLURI”
t.font=Font(name=“Arial”,bold=True,size=13,color=WHITE)
t.fill=PatternFill(“solid”,fgColor=DARK)
t.alignment=Alignment(horizontal=“center”,vertical=“center”)
ws.row_dimensions[1].height=30
hdrs=[“Month”,“Shifts”,“Total Hours”,“Gross Earnings(£)”,“Gym(£)”,“Other Deductions(£)”,“Est. Net Pay(£)”]
wdts=[16,10,14,20,10,22,18]
for col,(h,w) in enumerate(zip(hdrs,wdts),1):
c=ws.cell(row=2,column=col,value=h)
c.font=Font(name=“Arial”,bold=True,size=10,color=WHITE)
c.fill=PatternFill(“solid”,fgColor=GREEN)
c.alignment=Alignment(horizontal=“center”,vertical=“center”)
c.border=thin_border(WHITE)
ws.column_dimensions[get_column_letter(col)].width=w
ws.row_dimensions[2].height=22; ws.freeze_panes=“A3”
return ws

def update_summary(wb,month_name,shifts,total_hours,total_earnings):
ws=get_or_create_summary(wb); target_row=None
for row in range(3,ws.max_row+2):
val=ws.cell(row=row,column=1).value
if val==month_name: target_row=row; break
if val is None: target_row=row; break
gym=20.00; est_net=round(total_earnings-gym,2)
bg=LGREY if target_row%2==0 else WHITE
values=[month_name,shifts,total_hours,total_earnings,gym,“Enter Tax + NI”,est_net]
for col,val in enumerate(values,1):
c=ws.cell(row=target_row,column=col,value=val)
c.fill=PatternFill(“solid”,fgColor=bg)
c.font=Font(name=“Arial”,size=10,bold=(col in [1,4,7]),color=(“1A6B3C” if col==7 else “000000”))
c.alignment=Alignment(horizontal=“center”,vertical=“center”)
c.border=thin_border()
ws.row_dimensions[target_row].height=18

def get_month_totals(wb,month_name):
if month_name not in wb.sheetnames: return 0,0.0,0.0
ws=wb[month_name]; shifts=0; hours=0.0; earnings=0.0
for row in ws.iter_rows(min_row=5,values_only=True):
if row[0] and isinstance(row[0],int):
shifts+=1; hours+=float(row[10] or 0); earnings+=float(row[12] or 0)
return shifts,round(hours,1),round(earnings,2)

def save_to_excel(parsed,hrs,start_hr,end_hr,brk,rate,day_type,rate_period,earnings,dist,t_mins,wake_up,leave_home):
if os.path.exists(EXCEL_FILE): wb=openpyxl.load_workbook(EXCEL_FILE)
else:
wb=openpyxl.Workbook()
if “Sheet” in wb.sheetnames: del wb[“Sheet”]
get_or_create_summary(wb)
ws=get_or_create_month_sheet(wb,parsed[“date”])
next_row=max(ws.max_row+1,5); row_num=next_row-4; bg=LGRN if row_num%2==0 else WHITE
row_data=[row_num,parsed[“date”].strftime(”%d/%m/%Y”),parsed[“date”].strftime(”%A”),
parsed[“location”],parsed[“unit”],parsed[“shift_type”],day_type,
f”{start_hr:02d}:00”,f”{end_hr:02d}:00”,brk,hrs,rate,earnings,dist,t_mins,wake_up,leave_home]
for col,val in enumerate(row_data,1):
c=ws.cell(row=next_row,column=col,value=val)
c.fill=PatternFill(“solid”,fgColor=bg); c.font=Font(name=“Arial”,size=10)
c.alignment=Alignment(horizontal=“center”,vertical=“center”); c.border=thin_border()
ws.row_dimensions[next_row].height=18
sort_sheet_by_date(ws)
month_name=parsed[“date”].strftime(”%B %Y”)
shifts,total_hours,total_earnings=get_month_totals(wb,month_name)
update_summary(wb,month_name,shifts,total_hours,total_earnings)
if “Summary” in wb.sheetnames:
wb.move_sheet(“Summary”,offset=-wb.sheetnames.index(“Summary”))
wb.save(EXCEL_FILE)
return shifts,total_hours,total_earnings

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
if update.effective_user.id != YOUR_USER_ID: return
text=update.message.text.strip(); lower=text.lower()

```
if "confirmed shift" in lower or "new shift" in lower:
    await update.message.reply_text("Processing your shift... please wait.")
    try:
        parsed=parse_shift(text)
        work_lat,work_lng=geocode(parsed["geo_query"])
        dist=distance_km(HOME_LAT,HOME_LNG,work_lat,work_lng) if work_lat else 12.0
        t_mins=travel_time(dist)
        hrs,start_hr,end_hr,brk=shift_hours(parsed["shift_type"])
        rate,day_type,rate_period=get_pay_rate(parsed["date"],parsed["shift_type"])
        earnings=round(hrs*rate,2)
        wake_up,leave_home,shift_start_str=get_schedule(parsed["date"],start_hr,t_mins)
        total_shifts,total_hours,total_earnings=save_to_excel(
            parsed,hrs,start_hr,end_hr,brk,rate,day_type,rate_period,
            earnings,dist,t_mins,wake_up,leave_home)
        month_name=parsed["date"].strftime("%B %Y")
        day_name=parsed["date"].strftime("%A")
        msg=(f"SHIFT LOGGED\n{'='*30}\n\n"
             f"Date     : {parsed['date'].strftime('%d %B %Y')} ({day_name})\n"
             f"Location : {parsed['location']}\n"
             f"Unit     : {parsed['unit']}\n\n"
             f"SHIFT DETAILS\n"
             f"Type     : {parsed['shift_type']}\n"
             f"Start    : {start_hr:02d}:00\n"
             f"End      : {end_hr:02d}:00\n"
             f"Break    : {brk} hr\n"
             f"Hours    : {hrs} hrs\n\n"
             f"PAY ({rate_period})\n"
             f"Day Type : {day_type}\n"
             f"Rate     : £{rate:.2f}/hr\n"
             f"Earnings : £{earnings:.2f}\n\n"
             f"TRAVEL (Bus from {HOME_ADDRESS})\n"
             f"Distance : {dist} km\n"
             f"Bus Time : ~{t_mins} mins\n\n"
             f"YOUR SCHEDULE\n"
             f"Wake Up    : {wake_up}\n"
             f"Leave Home : {leave_home}\n"
             f"Arrive     : {shift_start_str}\n\n"
             f"{'='*30}\n"
             f"{month_name} TOTALS\n"
             f"Shifts: {total_shifts} | Hours: {total_hours} | £{total_earnings:.2f}\n"
             f"{'='*30}\n"
             f"Excel updated and sorted by date")
        await update.message.reply_text(msg)
    except Exception as e:
        await update.message.reply_text(f"Error: {str(e)}")

elif "good night" in lower:
    await update.message.reply_text(f"Good night, {YOUR_NAME}! Sleep well. Everything is tracked.")

elif "good morning" in lower:
    month_name=datetime.now().strftime("%B %Y")
    if os.path.exists(EXCEL_FILE):
        wb=openpyxl.load_workbook(EXCEL_FILE)
        shifts,hours,earnings=get_month_totals(wb,month_name)
    else: shifts,hours,earnings=0,0.0,0.0
    await update.message.reply_text(
        f"Good morning, {YOUR_NAME}!\n\n{month_name}:\nShifts: {shifts} | Hours: {hours} | Earned: £{earnings:.2f}\n\nHave a great shift!")

elif "summary" in lower or "total" in lower:
    month_name=datetime.now().strftime("%B %Y")
    if os.path.exists(EXCEL_FILE):
        wb=openpyxl.load_workbook(EXCEL_FILE)
        shifts,hours,earnings=get_month_totals(wb,month_name)
    else: shifts,hours,earnings=0,0.0,0.0
    await update.message.reply_text(
        f"{month_name} SUMMARY\n{'='*22}\nShifts: {shifts}\nHours: {hours}\nEarned: £{earnings:.2f}")

elif "help" in lower:
    await update.message.reply_text(
        f"Hi {YOUR_NAME}! Commands:\n\nPaste shift -> auto logged\n'summary' -> monthly totals\n'good morning' -> daily update\n'good night' -> sleep mode\n\nAuto: correct pay rate, distance, bus time, wake up time, monthly sheets, date sorting")
else:
    await update.message.reply_text(f"Hi {YOUR_NAME}! Send your shift confirmation and I'll log everything.\nType 'help' for commands.")
```

def main():
print(”=”*42)
print(”  SHIFT TRACKER BOT - LIVE”)
print(”=”*42)
print(f”  Name   : {YOUR_NAME}”)
print(f”  Home   : {HOME_ADDRESS}”)
print(f”  Rates  : Auto-switch 1 April 2026”)
print(”=”*42)
app=Application.builder().token(BOT_TOKEN).build()
app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
app.run_polling()

if **name** == “**main**”:
main()
